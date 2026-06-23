"""Example 5: Ad-Spend Reconciliation (non-coding, /goal). Carries reconciliation.xlsx.

DTC brand Lumen Goods reconciles Q2 ad invoices vs a messy bank export to a hard
gate (variance $0.00, 0 unmatched), with a human "ask-for-rule" step.

The waterfall is fully self-consistent: the Exceptions tab's variance_impact
column sums to the initial $1,402.88; resolving currency ($992.78), then the
duplicate ($321.70), then the 3 orphan charges ($88.40) walks the variance
1402.88 -> 410.10 -> 88.40 -> 0.00. The Matched tab balances both sides.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill

from _util import (
    EXAMPLES_DIR,
    cost_rows_from_cents,
    micros_to_usd,
    save_xlsx_deterministic,
    write_artifacts_index,
    write_csv,
    write_headline,
    write_text,
)

SLUG = "5-ad-spend-reconciliation"

# Matched truth: invoice (normalized USD) == bank charge. Balances to 7155.08.
MATCHED = [
    ("INV-1001", "Google Ads", 3250.00, "GOOGLE *ADS", "2026-04-04"),
    ("INV-1002", "Meta Ads", 2180.50, "FACEBK *ADS", "2026-04-06"),
    ("INV-1003", "TikTok Ads", 612.40, "TIKTOK ADS", "2026-04-10"),
    ("INV-1004", "LinkedIn Ads", 380.38, "LNKD *ADS", "2026-04-13"),
    ("INV-1005", "Reddit Ads", 410.10, "REDDIT ADS", "2026-04-19"),
    ("INV-1006", "Pinterest Ads", 321.70, "PINTEREST ADS", "2026-04-21"),
]
# Raw invoices (messy): two non-USD currencies, one duplicate, dates differ.
RAW_INVOICES = [
    ("INV-1001", "Google Ads", "2026-04-03", 3250.00, "USD"),
    ("INV-1002", "Meta Ads", "2026-04-05", 2180.50, "USD"),
    ("INV-1003", "TikTok Ads", "2026-04-09", 540.00, "EUR"),
    ("INV-1004", "LinkedIn Ads", "2026-04-12", 302.00, "GBP"),
    ("INV-1005", "Reddit Ads", "2026-04-18", 410.10, "USD"),
    ("INV-1006", "Pinterest Ads", "2026-04-20", 321.70, "USD"),
    ("INV-1006", "Pinterest Ads", "2026-04-20", 321.70, "USD"),  # DUPLICATE
]
# Orphan bank charges (no invoice) -> resolved to a fees/credit bucket.
ORPHANS = [
    ("2026-04-22", "FX CONVERSION FEE", 42.10),
    ("2026-04-23", "ADPLATFORM FEE", 34.30),
    ("2026-04-24", "AD CREDIT ADJUSTMENT", 12.00),
]
# Exceptions: variance_impact sums to the initial variance (1402.88).
EXCEPTIONS = [
    ("currency", "INV-1003 invoiced in EUR (540.00) - unmatched until normalized to $612.40", 612.40, "Normalized EUR->USD at the Q2 booked rate"),
    ("currency", "INV-1004 invoiced in GBP (302.00) - unmatched until normalized to $380.38", 380.38, "Normalized GBP->USD at the Q2 booked rate"),
    ("duplicate", "INV-1006 (Pinterest $321.70) ingested twice", 321.70, "De-duplicated; kept one"),
    ("orphan", "Bank 'FX CONVERSION FEE' $42.10 has no invoice", 42.10, "Classified as a bank fee (human rule)"),
    ("orphan", "Bank 'ADPLATFORM FEE' $34.30 has no invoice", 34.30, "Classified as a platform fee (human rule)"),
    ("orphan", "Bank 'AD CREDIT ADJUSTMENT' $12.00 has no invoice", 12.00, "Classified as a credit (human rule)"),
]
COST_CENTS = [60, 30, 40, 50, 20, 10]  # = 210c = $2.10
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(color="006100", bold=True)
RED_FONT = Font(color="9C0006", bold=True)


def generate() -> dict:
    d = EXAMPLES_DIR / SLUG
    d.mkdir(parents=True, exist_ok=True)
    (d / "inputs").mkdir(exist_ok=True)

    matched_total = round(sum(m[2] for m in MATCHED), 2)
    initial_variance = round(sum(e[2] for e in EXCEPTIONS), 2)  # 1402.88
    # Waterfall = variance remaining after resolving each category in order.
    cur = round(sum(e[2] for e in EXCEPTIONS if e[0] == "currency"), 2)
    dup = round(sum(e[2] for e in EXCEPTIONS if e[0] == "duplicate"), 2)
    orp = round(sum(e[2] for e in EXCEPTIONS if e[0] == "orphan"), 2)
    waterfall = [
        initial_variance,
        round(initial_variance - cur, 2),
        round(initial_variance - cur - dup, 2),
        round(initial_variance - cur - dup - orp, 2),
    ]

    # inputs/
    write_csv(
        d / "inputs" / "invoices.csv",
        ["invoice_id", "vendor", "invoice_date", "amount", "currency"],
        [[i[0], i[1], i[2], f"{i[3]:.2f}", i[4]] for i in RAW_INVOICES],
    )
    bank_rows = [[m[4], m[3], f"{m[2]:.2f}", "USD"] for m in MATCHED] + [[o[0], o[1], f"{o[2]:.2f}", "USD"] for o in ORPHANS]
    write_csv(d / "inputs" / "bank-export.csv", ["charge_date", "description", "amount", "currency"], bank_rows)

    # exceptions-resolved.csv
    write_csv(
        d / "exceptions-resolved.csv",
        ["category", "issue", "variance_impact_usd", "resolution"],
        [[e[0], e[1], f"{e[2]:.2f}", e[3]] for e in EXCEPTIONS],
    )

    # cost.csv
    rows, total_micros = cost_rows_from_cents(COST_CENTS)
    write_csv(
        d / "cost.csv",
        ["pass", "step", "input_tokens", "output_tokens", "cost_usd", "cumulative_usd"],
        [[i + 1, s, r["input_tokens"], r["output_tokens"], f'{r["cost_usd"]:.2f}', f'{r["cumulative_usd"]:.2f}']
         for i, (s, r) in enumerate(zip(["ingest", "normalize-currency", "dedup", "match", "ask-rule", "finalize"], rows))],
    )

    # loop-log.md (variance trajectory + the human ask-for-rule step)
    write_text(
        d / "loop-log.md",
        "# Loop log - Lumen Goods Q2 ad-spend reconciliation\n\n"
        "> Synthetic reconstruction for teaching. Validation-gate over messy data.\n\n"
        "## /goal\n"
        "Reconcile Q2 ad invoices against the bank export until **variance = $0.00 and "
        "0 unmatched rows**. If a charge cannot be classified, **ask for a rule** instead "
        "of guessing.\n\n"
        "## Variance trajectory\n"
        f"- Start: **${waterfall[0]:,.2f}** (currencies not normalized, a duplicate, 3 orphan charges)\n"
        f"- After normalizing EUR/GBP invoices: **${waterfall[1]:,.2f}** (-${cur:,.2f})\n"
        f"- After de-duplicating INV-1006: **${waterfall[2]:,.2f}** (-${dup:,.2f})\n"
        f"- After resolving 3 orphan charges: **${waterfall[3]:,.2f}** BALANCED (-${orp:,.2f})\n\n"
        "## Human ask-for-rule step\n"
        "The loop could not classify `AD CREDIT ADJUSTMENT` on its own. It **paused and "
        "asked**: \"Is a negative ad-credit a bank fee, a refund, or a contra-revenue line?\" "
        "Human rule: *treat ad credits as a credit to the fees bucket.* The loop applied the "
        "rule and re-ran to $0.00.\n",
    )

    _write_xlsx(d / "reconciliation.xlsx", matched_total, initial_variance, EXCEPTIONS)

    headline = {
        "example": 5,
        "title": "Ad-Spend Reconciliation",
        "start_variance_usd": waterfall[0],
        "end_variance_usd": waterfall[3],
        "waterfall_usd": waterfall,
        "waterfall_deltas_usd": [cur, dup, orp],
        "matched_total_usd": matched_total,
        "unmatched_rows": 0,
        "cost_usd": micros_to_usd(total_micros),
    }
    write_headline(d, headline)

    write_artifacts_index(
        d,
        "Example 5 - Ad-Spend Reconciliation",
        [
            ("Reconciliation workbook (.xlsx)", "reconciliation.xlsx"),
            ("Input: messy invoices", "inputs/invoices.csv"),
            ("Input: bank export", "inputs/bank-export.csv"),
            ("Exceptions resolved", "exceptions-resolved.csv"),
            ("Cost ledger (~$2.10)", "cost.csv"),
            ("Loop log (variance to $0.00)", "loop-log.md"),
            ("Headline numbers", "headline.json"),
        ],
        note="Synthetic reconstruction for teaching. Validation-gate over messy data; the Summary cell flips to $0.00 BALANCED.",
    )
    return headline


def _write_xlsx(path, matched_total, initial_variance, exceptions) -> None:
    wb = Workbook()
    raw = wb.active
    raw.title = "Raw"
    raw.append(["invoice_id", "vendor", "invoice_date", "amount", "currency"])
    for i in RAW_INVOICES:
        raw.append([i[0], i[1], i[2], i[3], i[4]])
    raw.append([])
    raw.append(["bank_charge_date", "description", "amount", "currency"])
    for m in MATCHED:
        raw.append([m[4], m[3], m[2], "USD"])
    for o in ORPHANS:
        raw.append([o[0], o[1], o[2], "USD"])

    norm = wb.create_sheet("Normalized")
    norm.append(["invoice_id", "vendor", "amount_usd"])
    for m in MATCHED:
        norm.append([m[0], m[1], m[2]])

    matched = wb.create_sheet("Matched")
    matched.append(["invoice_id", "vendor", "invoice_usd", "bank_description", "bank_usd", "bank_date"])
    for m in MATCHED:
        matched.append([m[0], m[1], m[2], m[3], m[2], m[4]])
    matched.append([])
    n = len(MATCHED)
    matched.append(["TOTAL", "", f"=SUM(C2:C{n + 1})", "", f"=SUM(E2:E{n + 1})", ""])

    exc = wb.create_sheet("Exceptions")
    exc.append(["category", "issue", "variance_impact_usd", "resolution"])
    for e in exceptions:
        exc.append([e[0], e[1], e[2], e[3]])
    exc.append(["TOTAL", "", f"=SUM(C2:C{len(exceptions) + 1})", ""])

    summ = wb.create_sheet("Summary")
    summ.append(["Metric", "Value"])
    summ.append(["Initial variance (USD)", initial_variance])
    summ.append(["Total resolved (USD)", initial_variance])
    summ.append(["Final variance (USD)", 0.00])
    summ.append(["Unmatched rows", 0])
    summ.append(["Status", "BALANCED"])
    summ["A1"].font = Font(bold=True)
    summ["B1"].font = Font(bold=True)
    # The headline cell: green when balanced (==0), red otherwise.
    summ["B4"].fill = GREEN_FILL
    summ["B4"].font = GREEN_FONT
    summ["B6"].fill = GREEN_FILL
    summ["B6"].font = GREEN_FONT
    summ.conditional_formatting.add(
        "B4",
        CellIsRule(operator="equal", formula=["0"], fill=GREEN_FILL, font=GREEN_FONT),
    )
    summ.conditional_formatting.add(
        "B4",
        CellIsRule(operator="notEqual", formula=["0"], font=RED_FONT),
    )
    for ws in (raw, norm, matched, exc, summ):
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 28

    save_xlsx_deterministic(wb, path)


if __name__ == "__main__":
    print(generate())
