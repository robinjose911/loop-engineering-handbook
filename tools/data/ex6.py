"""Example 6: RFP / Security Questionnaire Pack (non-coding, /goal). Carries ANSWER_PACK.xlsx.

Vendor Castora Analytics answers a 60-question questionnaire for buyer Meridian
Bank, drafting ONLY from an approved binder. Every answer must cite a source line
or it is auto-deleted and marked GAP. Pattern: citation claim-ledger (cite-or-cut).

The loop drafts "we're FedRAMP authorized," finds no source, self-deletes it, and
writes "GAP - needs human."
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from _util import (
    EXAMPLES_DIR,
    cost_rows_from_cents,
    micros_to_usd,
    save_xlsx_deterministic,
    time_seq,
    write_artifacts_index,
    write_csv,
    write_headline,
    write_jsonl,
    write_text,
)

SLUG = "6-rfp-questionnaire-pack"

# The only citable corpus. Each statement becomes one answered question.
CORPUS = {
    "SRC-SEC": ("security.md", "Information Security Policy", [
        "All data in transit is encrypted with TLS 1.3.",
        "Data at rest is encrypted with AES-256.",
        "We enforce SSO with SAML 2.0 for all employees.",
        "Multi-factor authentication is required for all production access.",
        "Secrets are stored in a managed vault, never in source control.",
        "We run quarterly third-party penetration tests.",
        "Vulnerabilities are triaged within one business day.",
        "Production access is logged and reviewed monthly.",
        "We follow a documented secure-SDLC with mandatory code review.",
        "Endpoint disk encryption is enforced on all company laptops.",
    ]),
    "SRC-INFRA": ("infrastructure.md", "Infrastructure & Availability", [
        "Production runs in three availability zones in a single region.",
        "We target 99.9% monthly uptime.",
        "Backups are taken every 6 hours and tested monthly.",
        "Infrastructure is provisioned as code and peer-reviewed.",
        "We use a CDN with DDoS protection at the edge.",
        "Autoscaling is configured for all stateless services.",
        "Database failover is automated with a standby replica.",
        "All servers are patched on a 30-day cadence.",
        "Logs are centralized and retained for one year.",
        "Capacity is reviewed weekly against demand forecasts.",
    ]),
    "SRC-DATA": ("data-handling.md", "Data Handling & Privacy", [
        "Customer data is logically segregated per tenant.",
        "We support data residency in the EU and the US.",
        "Personal data is deleted within 30 days of account closure.",
        "We maintain a record of processing activities under GDPR.",
        "Data subject access requests are fulfilled within 30 days.",
        "We do not sell customer data to third parties.",
        "PII is masked in non-production environments.",
        "Data retention windows are configurable per tenant.",
        "Encryption keys are rotated annually.",
        "Customers can export their data on request.",
    ]),
    "SRC-COMP": ("compliance.md", "Compliance & Certifications", [
        "We hold a current SOC 2 Type II report.",
        "We are ISO 27001 certified.",
        "Our SOC 2 report is available under NDA.",
        "We undergo an annual external compliance audit.",
        "We maintain GDPR and CCPA compliance programs.",
        "A current subprocessor list is published and versioned.",
        "Penetration-test summaries are shared under NDA.",
        "We map controls to the CIS Critical Security Controls.",
    ]),
    "SRC-BCP": ("business-continuity.md", "Business Continuity & IR", [
        "We maintain a documented incident-response plan.",
        "Customers are notified of confirmed breaches within 72 hours.",
        "Our RTO is 4 hours and our RPO is 6 hours.",
        "We run an annual disaster-recovery exercise.",
        "On-call coverage is 24/7 for production incidents.",
        "Post-incident reviews are blameless and documented.",
        "Backups are stored in a region separate from production.",
        "We operate a public status page for incident communication.",
    ]),
    "SRC-ACCESS": ("access-control.md", "Access Control & HR", [
        "Access follows least-privilege and is role-based.",
        "Access is reviewed and recertified every quarter.",
        "Employees complete security training at onboarding and annually.",
        "Background checks are performed on all employees.",
        "Access is revoked within 24 hours of termination.",
        "Privileged actions require a second approver.",
        "Contractors are subject to the same access policies as employees.",
        "Shared accounts are prohibited.",
        "Session timeouts are enforced after 15 minutes idle.",
        "We maintain an asset inventory reviewed semi-annually.",
        "Just-in-time access is used for sensitive production tasks.",
        "All access changes are recorded in an audit log.",
    ]),
}
# Two questions with no support in the binder -> drafted, unsupported, cut to GAP.
GAP_QUESTIONS = [
    ("Are you FedRAMP authorized?", "we are FedRAMP authorized", "No source in the binder; the draft claim was self-deleted."),
    ("Do you carry a $10M cyber-insurance policy?", "we carry $10M cyber-insurance", "No source in the binder; needs a human to confirm coverage."),
]
COST_CENTS = [40] + [3] * 60 + [25]  # draft + 60 cite-checks + finalize = $2.45


def generate() -> dict:
    d = EXAMPLES_DIR / SLUG
    d.mkdir(parents=True, exist_ok=True)
    sot = d / "inputs" / "source_of_truth"
    sot.mkdir(parents=True, exist_ok=True)

    # Write the citable corpus.
    for sid, (fname, title, statements) in CORPUS.items():
        body = "\n".join(f"- {s}" for s in statements)
        write_text(sot / fname, f"# {title}\n\nSource id: `{sid}`\n\n{body}\n")

    # Build answered rows from the corpus + the 2 GAP rows.
    answered = []
    qid = 0
    for sid, (fname, title, statements) in CORPUS.items():
        for s in statements:
            qid += 1
            confidence = "high" if qid % 5 else "medium"
            answered.append({
                "question_id": f"Q{qid:02d}",
                "question": f"[{title}] Confirm: {s}",
                "answer": f"Yes. {s}",
                "source_id": sid,
                "source_quote": s,
                "confidence": confidence,
                "status": "ANSWERED",
            })
    gaps = []
    for gq, _claim, why in GAP_QUESTIONS:
        qid += 1
        gaps.append({
            "question_id": f"Q{qid:02d}",
            "question": gq,
            "answer": "GAP - needs human",
            "source_id": "",
            "source_quote": "",
            "confidence": "none",
            "status": "GAP",
        })
    rows = answered + gaps
    total_q = len(rows)
    answered_n = len(answered)
    gap_n = len(gaps)

    _write_answer_pack(d / "ANSWER_PACK.xlsx", rows)

    # claim_ledger.csv: per-claim PASS/FAIL (cite-or-cut).
    ledger = [[r["question_id"], r["question"][:60], "yes" if r["source_id"] else "no",
               "PASS" if r["status"] == "ANSWERED" else "FAIL"] for r in rows]
    write_csv(d / "claim_ledger.csv", ["question_id", "claim", "has_source", "verdict"], ledger)

    # loop-log.jsonl: includes the self-deleted FedRAMP claim.
    ts = time_seq("2026-06-02T11:00:00", 45)
    records = [{"turn": 1, "ts": next(ts), "status": "drafting", "note": "drafting answers from the approved binder only"}]
    turn = 1
    for gq, claim, why in GAP_QUESTIONS:
        turn += 1
        records.append({
            "turn": turn,
            "ts": next(ts),
            "status": "gap_marked",
            "question": gq,
            "drafted_claim": claim,
            "action": "no citation found -> claim self-deleted -> marked GAP",
            "note": why,
        })
    write_jsonl(d / "loop-log.jsonl", records)

    # progress.md
    write_text(
        d / "progress.md",
        "# Progress - Castora Analytics questionnaire for Meridian Bank\n\n"
        "> Synthetic reconstruction for teaching. Citation claim-ledger (cite-or-cut).\n\n"
        f"- Questions: **{total_q}**\n"
        f"- Answered (with a cited source line): **{answered_n}/{total_q}**\n"
        f"- GAPs (no source - escalated to a human): **{gap_n}**\n\n"
        "The loop drafted \"we're FedRAMP authorized,\" found no source line in the binder, "
        "**self-deleted the claim**, and wrote \"GAP - needs human.\" Cite-or-cut, enforced "
        "on every answer.\n",
    )

    rows_cost, total_micros = cost_rows_from_cents(COST_CENTS)
    write_csv(
        d / "cost.csv",
        ["pass", "input_tokens", "output_tokens", "cost_usd", "cumulative_usd"],
        [[i + 1, r["input_tokens"], r["output_tokens"], f'{r["cost_usd"]:.2f}', f'{r["cumulative_usd"]:.2f}']
         for i, r in enumerate(rows_cost)],
    )

    headline = {
        "example": 6,
        "title": "RFP / Security Questionnaire Pack",
        "total_questions": total_q,
        "answered": answered_n,
        "gaps": gap_n,
        "sources": len(CORPUS),
        "coverage_pct": round(100 * answered_n / total_q, 1),
        "cost_usd": micros_to_usd(total_micros),
    }
    write_headline(d, headline)

    write_artifacts_index(
        d,
        "Example 6 - RFP / Security Questionnaire Pack",
        [
            ("Answer pack (.xlsx)", "ANSWER_PACK.xlsx"),
            ("Source of truth (binder)", "inputs/source_of_truth/security.md"),
            ("Claim ledger (PASS/FAIL)", "claim_ledger.csv"),
            ("Loop log (self-deleted claim)", "loop-log.jsonl"),
            ("Progress", "progress.md"),
            ("Cost ledger", "cost.csv"),
            ("Headline numbers", "headline.json"),
        ],
        note="Synthetic reconstruction for teaching. Cite-or-cut: no answer ships without a source line.",
    )
    return headline


def _write_answer_pack(path, rows) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Answers"
    header = ["question_id", "question", "answer", "source_id", "source_quote", "confidence", "status"]
    ws.append(header)
    for c in range(1, len(header) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    gap_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for r in rows:
        ws.append([r[h] for h in header])
        if r["status"] == "GAP":
            ws.cell(row=ws.max_row, column=7).fill = gap_fill
    widths = {"A": 10, "B": 44, "C": 40, "D": 12, "E": 44, "F": 12, "G": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    save_xlsx_deterministic(wb, path)


if __name__ == "__main__":
    print(generate())
