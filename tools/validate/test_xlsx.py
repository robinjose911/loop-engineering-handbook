"""Unit test for the two real .xlsx artifacts (Stage 1.3).

reconciliation.xlsx: Summary variance == 0.00 and unmatched == 0; the Matched tab
balances; the Exceptions impacts sum to the initial variance. ANSWER_PACK.xlsx:
every ANSWERED row cites a source; exactly 2 GAP rows; coverage matches progress.
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
EXAMPLES = REPO_ROOT / "examples"


def _headline(slug: str) -> dict:
    return json.loads((EXAMPLES / slug / "headline.json").read_text())


def test_reconciliation_xlsx_balances():
    slug = "5-ad-spend-reconciliation"
    h = _headline(slug)
    wb = load_workbook(EXAMPLES / slug / "reconciliation.xlsx", data_only=False)
    assert set(wb.sheetnames) == {"Raw", "Normalized", "Matched", "Exceptions", "Summary"}

    # Summary: variance 0.00, unmatched 0.
    summ = wb["Summary"]
    labels = {summ.cell(r, 1).value: summ.cell(r, 2).value for r in range(1, summ.max_row + 1)}
    assert float(labels["Final variance (USD)"]) == pytest.approx(0.0, abs=0.005)
    assert int(labels["Unmatched rows"]) == 0
    assert labels["Status"] == "BALANCED"

    # Matched tab balances: sum(invoice_usd) == sum(bank_usd).
    matched = wb["Matched"]
    inv = bank = 0.0
    for r in range(2, matched.max_row + 1):
        if str(matched.cell(r, 1).value or "").startswith("INV-"):
            inv += float(matched.cell(r, 3).value)
            bank += float(matched.cell(r, 5).value)
    assert inv == pytest.approx(bank, abs=0.005)
    assert inv == pytest.approx(h["matched_total_usd"], abs=0.005)

    # Exceptions impacts sum to the initial variance.
    exc = wb["Exceptions"]
    impact = sum(
        float(exc.cell(r, 3).value)
        for r in range(2, exc.max_row + 1)
        if exc.cell(r, 1).value in ("currency", "duplicate", "orphan")
    )
    assert impact == pytest.approx(1402.88, abs=0.005)
    assert h["start_variance_usd"] == 1402.88


def test_answer_pack_xlsx_cite_or_cut():
    slug = "6-rfp-questionnaire-pack"
    h = _headline(slug)
    wb = load_workbook(EXAMPLES / slug / "ANSWER_PACK.xlsx")
    ws = wb["Answers"]
    header = [c.value for c in ws[1]]
    assert header == ["question_id", "question", "answer", "source_id", "source_quote", "confidence", "status"]

    answered = gaps = 0
    for r in range(2, ws.max_row + 1):
        row = {header[c]: ws.cell(r, c + 1).value for c in range(len(header))}
        if row["status"] == "ANSWERED":
            answered += 1
            assert row["source_id"], f"ANSWERED row {r} missing source_id"
            assert row["source_quote"], f"ANSWERED row {r} missing source_quote"
        elif row["status"] == "GAP":
            gaps += 1
            assert not row["source_id"], f"GAP row {r} should not cite a source"

    assert gaps == 2, f"expected exactly 2 GAP rows, got {gaps}"
    assert answered == h["answered"] == 58
    assert answered + gaps == h["total_questions"] == 60

    # Independent cross-check against the narrative (literal, not derived from
    # the workbook we are validating).
    progress = (EXAMPLES / slug / "progress.md").read_text()
    assert "58/60" in progress
